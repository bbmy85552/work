#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def process_merged_cells(file_path):
    """
    处理包含合并单元格的Excel文件
    """
    # 使用openpyxl加载工作簿
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # 获取所有合并单元格的范围
    merged_ranges = ws.merged_cells.ranges

    # 创建一个字典来存储合并单元格的值
    merged_values = {}

    # 处理每个合并单元格范围
    for merged_range in merged_ranges:
        # 获取合并区域的左上角单元格的值
        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left_cell.value

        # 将值填充到合并区域的所有单元格
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value

    # 创建一个新的工作表来处理数据
    data = []
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            # 首先检查是否在合并单元格字典中
            if (row, col) in merged_values:
                row_data.append(merged_values[(row, col)])
            else:
                # 否则获取原始单元格的值
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
        data.append(row_data)

    # 获取第一行作为列名
    columns = data[0] if data else []
    # 获取数据行
    rows = data[1:] if len(data) > 1 else []

    return pd.DataFrame(rows, columns=columns)

def convert_excel_to_csv():
    """
    将Excel文件转换为CSV格式
    """
    # 文件路径
    variables_file = "变量表.xlsx"
    data_file = "数据(1).xlsx"

    print("开始处理Excel文件...")

    # 处理变量表.xlsx（包含合并单元格）
    print(f"正在处理 {variables_file}...")
    try:
        df_variables = process_merged_cells(variables_file)

        # 清理列名（去除可能的空格和换行符）
        df_variables.columns = [str(col).strip() if col is not None else f"Column_{i}"
                               for i, col in enumerate(df_variables.columns)]

        # 保存为CSV
        output_csv1 = "变量表.csv"
        df_variables.to_csv(output_csv1, index=False, encoding='utf-8-sig')
        print(f"✓ {variables_file} 已成功转换为 {output_csv1}")
        print(f"  - 行数: {len(df_variables)}")
        print(f"  - 列数: {len(df_variables.columns)}")

    except Exception as e:
        print(f"✗ 处理 {variables_file} 时出错: {e}")

    # 处理数据(1).xlsx（正常Excel文件）
    print(f"\n正在处理 {data_file}...")
    try:
        # 使用pandas直接读取
        df_data = pd.read_excel(data_file)

        # 清理列名
        df_data.columns = [str(col).strip() if col is not None else f"Column_{i}"
                          for i, col in enumerate(df_data.columns)]

        # 保存为CSV
        output_csv2 = "数据(1).csv"
        df_data.to_csv(output_csv2, index=False, encoding='utf-8-sig')
        print(f"✓ {data_file} 已成功转换为 {output_csv2}")
        print(f"  - 行数: {len(df_data)}")
        print(f"  - 列数: {len(df_data.columns)}")

    except Exception as e:
        print(f"✗ 处理 {data_file} 时出错: {e}")

    print("\n所有文件转换完成！")

if __name__ == "__main__":
    convert_excel_to_csv()